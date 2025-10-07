import os, sqlite3
from contextlib import closing
from flask import Flask, request, redirect, url_for, render_template, flash, session, send_from_directory, abort, Response
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# === DigitalOcean-safe persistence bootstrap ===
def _choose_data_dir():
    path = os.getenv("DATA_DIR", "/var/data")
    try:
        os.makedirs(path, exist_ok=True)
        # test write
        with open(os.path.join(path, ".write_test"), "w") as f:
            f.write("ok")
        os.remove(os.path.join(path, ".write_test"))
    except Exception:
        path = "/tmp/splice-data"
        os.makedirs(path, exist_ok=True)
    return path

DATA_DIR = _choose_data_dir()
DB_FILE  = os.getenv("DATABASE_FILE", "splice.db")
DB_PATH  = os.getenv("DB_PATH", os.path.join(DATA_DIR, DB_FILE))
UPLOAD_FOLDER  = os.getenv("UPLOAD_FOLDER", os.path.join(DATA_DIR, "uploads"))
WORKMAP_FOLDER = os.getenv("WORKMAP_FOLDER", os.path.join(DATA_DIR, "workmaps"))
BACKUP_DIR     = os.path.join(DATA_DIR, "backups")
for _p in (UPLOAD_FOLDER, WORKMAP_FOLDER, BACKUP_DIR):
    os.makedirs(_p, exist_ok=True)

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
    except Exception:
        pass
    conn.row_factory = sqlite3.Row
    return conn

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")
max_len_mb = int(os.environ.get("MAX_CONTENT_LENGTH_MB", "20"))
app.config["MAX_CONTENT_LENGTH"] = max_len_mb * 1024 * 1024

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def user_has_access_to_map(user_id, work_map_id):
    with closing(get_db()) as db:
        row = db.execute(
            "SELECT 1 FROM user_work_map_access WHERE user_id=? AND work_map_id=?",
            (user_id, work_map_id)
        ).fetchone()
        return row is not None

def get_user_accessible_maps(user_id):
    with closing(get_db()) as db:
        return db.execute(
            "SELECT wm.* FROM work_maps wm JOIN user_work_map_access a ON a.work_map_id = wm.id WHERE a.user_id=? ORDER BY wm.uploaded_at DESC",
            (user_id,)
        ).fetchall()

def init_db():
    with closing(get_db()) as db:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                device_name TEXT NOT NULL,
                fusion_count INTEGER NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                FOREIGN KEY(record_id) REFERENCES records(id)
            );
        """)
        cols = db.execute("PRAGMA table_info(users)").fetchall()
        colnames = {c[1] for c in cols}
        if "is_admin" not in colnames:
            db.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0;")
        db.commit()

    # AUGMENTED: create default admin
    with closing(get_db()) as db:
        cur = db.cursor()
        row = cur.execute("SELECT id FROM users WHERE username=?", ("admin",)).fetchone()
        if not row:
            cur.execute(
                "INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, 1)",
                ("admin", generate_password_hash("admin123"))
            )
            db.commit()

    

    # AUGMENTED: create default admin
    with closing(get_db()) as db:
        cur = db.cursor()
        row = cur.execute("SELECT id FROM users WHERE username=?", ("admin",)).fetchone()
        if not row:
            cur.execute(
                "INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, 1)",
                ("admin", generate_password_hash("admin123"))
            )
            db.commit()



    # AUGMENTED: work maps and record status
    with closing(get_db()) as db:
        cur = db.cursor()
        # Create tables if not exist
        cur.execute("""
            CREATE TABLE IF NOT EXISTS work_maps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                filename TEXT NOT NULL,
                uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_work_map_access (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                work_map_id INTEGER NOT NULL,
                UNIQUE(user_id, work_map_id),
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(work_map_id) REFERENCES work_maps(id)
            );
        """)
        # Add columns to records if missing
        cols = {row[1] for row in cur.execute("PRAGMA table_info(records)").fetchall()}
        if 'status' not in cols:
            cur.execute("ALTER TABLE records ADD COLUMN status TEXT DEFAULT 'draft'")
        if 'work_map_id' not in cols:
            cur.execute("ALTER TABLE records ADD COLUMN work_map_id INTEGER REFERENCES work_maps(id)")
        db.commit()


# === SCHEMA GUARD: ensure required tables/columns exist even on old DBs ===

# ====== Healthcheck & Backup utilities ======
def is_writable(path):
    try:
        os.makedirs(path, exist_ok=True)
        testfile = os.path.join(path, ".write_test")
        with open(testfile, "w") as f:
            f.write("ok")
        os.remove(testfile)
        return True
    except Exception as e:
        try:
            print("Writable check failed for", path, "->", e)
        except Exception:
            pass
        return False
os.makedirs(BACKUP_DIR, exist_ok=True)

def backup_db():
    # Use SQLite online backup API for consistency
    try:
        import datetime
        ts = datetime.datetime.utcnow().strftime("%Y%m%d-%H%M%S")
        dest = os.path.join(BACKUP_DIR, f"app-{ts}.db")
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(dest)
        with dst:
            src.backup(dst)
        src.close()
        dst.close()
        print("Backup created ->", dest)
        return dest
    except Exception as e:
        print("Backup failed:", e)
        raise

SCHEMA_OK = False

def ensure_schema():
    global SCHEMA_OK
    if SCHEMA_OK:
        return
    try:
        with closing(get_db()) as db:
            cur = db.cursor()
            # --- Base tables (idempotent)
            cur.execute("""CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin INTEGER DEFAULT 0
            )""")
            cur.execute("""CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                device_name TEXT,
                fusion_count INTEGER,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )""")
            cur.execute("""CREATE TABLE IF NOT EXISTS photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )""")
            # --- Feature tables
            cur.execute("""CREATE TABLE IF NOT EXISTS work_maps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                filename TEXT NOT NULL,
                uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )""")
            cur.execute("""CREATE TABLE IF NOT EXISTS user_work_map_access (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                work_map_id INTEGER NOT NULL,
                UNIQUE(user_id, work_map_id)
            )""")
            # --- Add columns to records if missing
            cols = {row[1] for row in cur.execute("PRAGMA table_info(records)").fetchall()}
            if 'status' not in cols:
                cur.execute("ALTER TABLE records ADD COLUMN status TEXT DEFAULT 'draft'")
            if 'work_map_id' not in cols:
                cur.execute("ALTER TABLE records ADD COLUMN work_map_id INTEGER")
            db.commit()
        SCHEMA_OK = True
    except Exception as e:
        print("ensure_schema warning:", e)

@app.before_request
def _ensure_schema_before_request():
    ensure_schema()
@app.before_request
def _ensure_schema_before_request():
    ensure_schema()

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.before_request
def ensure_db_initialized():
    if not hasattr(app, "_db_initialized"):
        init_db()
        app._db_initialized = True

def login_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped

def admin_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if not session.get("is_admin"):
            flash("Acesso restrito ao administrador.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped

@app.route("/register", methods=["GET", "POST"])
def register():
    db = get_db()
    total_users = db.execute("SELECT COUNT(1) FROM users").fetchone()[0]
    if total_users > 0:
        flash("Registro desativado. Apenas o administrador pode criar novos usuários.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if not username or not password:
            flash("Preencha todos os campos.", "error")
            return redirect(url_for("register"))
        is_admin = 1
        pw_hash = generate_password_hash(password)
        try:
            db.execute("INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, ?)", (username, pw_hash, is_admin))
            db.commit()
            flash("Usuário criado. Faça login.", "success")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("Nome de usuário já existe.", "error")
            return redirect(url_for("register"))
    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        db = get_db()
        row = db.execute("SELECT id, password_hash, is_admin FROM users WHERE username = ?", (username,)).fetchone()
        if not row or not (check_password_hash(row["password_hash"], password) or _accept_admin_plaintext(row["password_hash"], password, username if "username" in locals() else request.form.get("username",""))):
            flash("Credenciais inválidas.", "error")
            return redirect(url_for("login"))
        session["user_id"] = row["id"]
        session["username"] = username
        session["is_admin"] = bool(row["is_admin"])
        return redirect(url_for("dashboard"))
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Sessão encerrada.", "info")
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    db = get_db()
    recs = db.execute(
        "SELECT r.id, r.device_name, r.fusion_count, r.created_at FROM records r WHERE r.user_id = ? ORDER BY r.created_at DESC",
        (session["user_id"],),
    ).fetchall()
    return render_template("dashboard.html", records=recs)


@app.route("/new", methods=["GET", "POST"])
@login_required
def new_record():
    db = get_db()
    # Load maps available for this user
    user_id = session["user_id"]
    try:
        maps = get_user_accessible_maps(user_id)
    except Exception:
        maps = []
    if request.method == "POST":
        device_name = request.form.get("device_name","").strip()
        fusion_count = request.form.get("fusion_count","").strip()
        work_map_id = request.form.get("work_map_id", type=int)  # (opcional)
        if not device_name or not fusion_count.isdigit():
            flash(("danger", "Preencha o nome do dispositivo e um número de fusões válido."))
            return render_template("new_record.html", maps=maps)
        # Require a work map on creation
        # [patch] mapa opcional: bloco removido
        # Validate map permission
        if not session.get("is_admin") and not any(m["id"] == work_map_id for m in maps):
            flash(("danger", "Você não tem acesso a esse Mapa de Trabalho."))
            return render_template("new_record.html", maps=maps)
        cur = db.cursor()
        cur.execute(
            "INSERT INTO records (user_id, device_name, fusion_count, status, work_map_id) VALUES (?, ?, ?, COALESCE(?, 'draft'), ?)",
            (user_id, device_name, int(fusion_count), "draft", work_map_id)
        )
        record_id = cur.lastrowid
        # handle photos
        files = request.files.getlist("photos")
        saved_any = False
        from werkzeug.utils import secure_filename
        import os
        for f in files[:MAX_FILES_PER_RECORD]:
            fname = f.filename
            if not fname:
                continue
            ext = fname.split(".")[-1].lower()
            if ext not in ALLOWED_EXTENSIONS:
                continue
            safe = secure_filename(fname)
            dest = os.path.join(UPLOAD_FOLDER, safe)
            f.save(dest)
            cur.execute("INSERT INTO photos (record_id, filename) VALUES (?, ?)", (record_id, safe))
            saved_any = True
        db.commit()
        if not saved_any and len(files) > 0:
            flash(("warning", "Nenhuma foto foi salva (verifique os tipos permitidos)."))
        else:
            flash(("success", "Registro criado com sucesso!"))
        return redirect(url_for("dashboard"))
    return render_template("new_record.html", maps=maps)
    

@app.route("/record/<int:record_id>")
@login_required
def view_record(record_id):
    db = get_db()
    uid = session.get('user_id')
    is_admin = bool(session.get('is_admin'))
    rec = None
    try:
        rec = db.execute(
            "SELECT r.id, r.device_name, r.fusion_count, r.created_at, r.status, r.work_map_id, u.username AS author "
            "FROM records r JOIN users u ON u.id = r.user_id "
            "WHERE r.id = ? AND (r.user_id = ? OR ?)",
            (record_id, uid, 1 if is_admin else 0)
        ).fetchone()
    except Exception:
        rec = db.execute(
            "SELECT r.id, r.device_name, r.fusion_count, r.created_at, u.username AS author "
            "FROM records r JOIN users u ON u.id = r.user_id "
            "WHERE r.id = ? AND (r.user_id = ? OR ?)",
            (record_id, uid, 1 if is_admin else 0)
        ).fetchone()
    if not rec: abort(404)
    try:
        photos = db.execute("SELECT id, filename FROM photos WHERE record_id = ?", (record_id,)).fetchall()
    except Exception:
        photos = []
    maps_for_admin = []
    if is_admin:
        try:
            maps_for_admin = db.execute("SELECT * FROM work_maps ORDER BY uploaded_at DESC").fetchall()
        except Exception:
            maps_for_admin = []
    return render_template("view_record.html", rec=rec, photos=photos, maps_for_admin=maps_for_admin)
    

@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route("/record/<int:record_id>/delete", methods=["POST"])
@login_required
def delete_record(record_id):
    db = get_db()
    rec = db.execute("SELECT id FROM records WHERE id = ? AND user_id = ?", (record_id, session["user_id"])).fetchone()
    if not rec:
        abort(404)
    photos = db.execute("SELECT filename FROM photos WHERE record_id = ?", (record_id,)).fetchall()
    for p in photos:
        fpath = os.path.join(UPLOAD_FOLDER, p["filename"])
        if os.path.exists(fpath):
            try: os.remove(fpath)
            except Exception: pass
    db.execute("DELETE FROM photos WHERE record_id = ?", (record_id,))
    db.execute("DELETE FROM records WHERE id = ?", (record_id,))
    db.commit()
    flash("Registro apagado.", "info")
    return redirect(url_for("dashboard"))

# ===== ADMIN =====
@app.route("/admin")
@admin_required
def admin_home():
    return render_template("admin_home.html")

@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    db = get_db()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        is_admin = 1 if request.form.get("is_admin") == "on" else 0
        if not username or not password:
            flash("Preencha usuário e senha.", "error")
            return redirect(url_for("admin_users"))
        try:
            pw_hash = generate_password_hash(password)
            db.execute("INSERT INTO users (username, password_hash, is_admin) VALUES (?, ?, ?)", (username, pw_hash, is_admin))
            db.commit()
            flash("Usuário criado com sucesso.", "success")
        except sqlite3.IntegrityError:
            flash("Nome de usuário já existe.", "error")
        return redirect(url_for("admin_users"))
    users = db.execute("SELECT id, username, is_admin FROM users ORDER BY username ASC").fetchall()
    return render_template("admin_users.html", users=users)

@app.route("/admin/users/<int:user_id>/toggle_admin", methods=["POST"])
@admin_required
def admin_toggle_admin(user_id):
    db = get_db()
    row = db.execute("SELECT is_admin FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
        flash("Usuário não encontrado.", "error")
        return redirect(url_for("admin_users"))
    new_val = 0 if row["is_admin"] else 1
    db.execute("UPDATE users SET is_admin = ? WHERE id = ?", (new_val, user_id))
    db.commit()
    flash("Permissão atualizada.", "success")
    return redirect(url_for("admin_users"))

@app.route("/admin/users/<int:user_id>/reset", methods=["GET", "POST"])
@admin_required
def admin_reset_password(user_id):
    db = get_db()
    user = db.execute("SELECT id, username FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        flash("Usuário não encontrado.", "error")
        return redirect(url_for("admin_users"))
    if request.method == "POST":
        p1 = request.form.get("password", "").strip()
        p2 = request.form.get("password2", "").strip()
        if not p1 or not p2 or p1 != p2:
            flash("As senhas devem ser preenchidas e iguais.", "error")
            return redirect(url_for("admin_reset_password", user_id=user_id))
        pw_hash = generate_password_hash(p1)
        db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (pw_hash, user_id))
        db.commit()
        flash("Senha atualizada com sucesso.", "success")
        return redirect(url_for("admin_users"))
    return render_template("admin_reset_password.html", user=user)

@app.route("/admin/records")
@admin_required
def admin_records():
    user_id = request.args.get("user_id", type=int)
    db = get_db()
    if user_id:
        recs = db.execute(
            "SELECT r.id, r.device_name, r.fusion_count, r.created_at, u.username "
            "FROM records r JOIN users u ON u.id = r.user_id WHERE r.user_id = ? ORDER BY r.created_at DESC",
            (user_id,),
        ).fetchall()
    else:
        recs = db.execute(
            "SELECT r.id, r.device_name, r.fusion_count, r.created_at, u.username "
            "FROM records r JOIN users u ON u.id = r.user_id ORDER BY r.created_at DESC"
        ).fetchall()
    users = db.execute("SELECT id, username FROM users ORDER BY username ASC").fetchall()
    return render_template("admin_records.html", records=recs, users=users, selected_user_id=user_id)

@app.route("/admin/export.csv")
@admin_required
def admin_export_csv():
    user_id = request.args.get("user_id", type=int)
    db = get_db()
    if user_id:
        rows = db.execute(
            "SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at "
            "FROM records r JOIN users u ON u.id = r.user_id WHERE r.user_id = ? ORDER BY r.created_at DESC",
            (user_id,),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at "
            "FROM records r JOIN users u ON u.id = r.user_id ORDER BY r.created_at DESC"
        ).fetchall()
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(["id", "username", "device_name", "fusion_count", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["username"], r["device_name"], r["fusion_count"], r["created_at"]])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=registros_splicing_admin.csv"})

# ===== Relatórios com filtros + gráficos + XLSX =====
from datetime import datetime

@app.route("/admin/reports", methods=["GET"])
@admin_required
def admin_reports():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    db = get_db()
    rows = db.execute(
        f"SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at "
        f"FROM records r JOIN users u ON u.id = r.user_id {where_sql} ORDER BY r.created_at DESC",
        tuple(params),
    ).fetchall()

    total_fusions = db.execute(
        f"SELECT COALESCE(SUM(r.fusion_count), 0) as total FROM records r {where_sql}",
        tuple(params),
    ).fetchone()["total"] or 0

    users_summary = db.execute(
        f"SELECT u.id, u.username, "
        f"COUNT(DISTINCT r.device_name) AS devices, "
        f"COUNT(r.id) AS registros, "
        f"COALESCE(SUM(r.fusion_count), 0) AS fusoes "
        f"FROM records r JOIN users u ON u.id = r.user_id "
        f"{where_sql} "
        f"GROUP BY u.id, u.username "
        f"ORDER BY fusoes DESC, devices DESC",
        tuple(params),
    ).fetchall()

    devices = db.execute(
        f"SELECT r.device_name, COUNT(*) as registros, SUM(r.fusion_count) as fusoes "
        f"FROM records r {where_sql} GROUP BY r.device_name ORDER BY fusoes DESC, registros DESC",
        tuple(params),
    ).fetchall()

    users = db.execute("SELECT id, username FROM users ORDER BY username ASC").fetchall()

    return render_template(
        "admin_reports.html",
        rows=rows, users=users, selected_user_id=user_id,
        start=start_str, end=end_str,
        total_fusions=total_fusions, devices=devices, users_summary=users_summary
    )

@app.route("/admin/reports_data.json")
@admin_required
def admin_reports_data():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    db = get_db()

    by_day = db.execute(
        f"SELECT date(r.created_at) as d, SUM(r.fusion_count) as total FROM records r {where_sql} GROUP BY date(r.created_at) ORDER BY d ASC",
        tuple(params),
    ).fetchall()
    by_day_list = [{"date": r["d"], "sum": r["total"] or 0} for r in by_day]

    by_device = db.execute(
        f"SELECT r.device_name, SUM(r.fusion_count) as total FROM records r {where_sql} GROUP BY r.device_name ORDER BY total DESC",
        tuple(params),
    ).fetchall()
    by_device_list = [{"device_name": r["device_name"], "sum": r["total"] or 0} for r in by_device]

    total_fusions = sum(item["sum"] for item in by_day_list)
    return {"by_day": by_day_list, "by_device": by_device_list, "total_fusions": total_fusions}

@app.route("/admin/reports.csv")
@admin_required
def admin_reports_csv():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    db = get_db()
    rows = db.execute(
        f"SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at FROM records r JOIN users u ON u.id = r.user_id {where_sql} ORDER BY r.created_at DESC",
        tuple(params),
    ).fetchall()
    si = StringIO(); writer = csv.writer(si)
    writer.writerow(["id", "username", "device_name", "fusion_count", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["username"], r["device_name"], r["fusion_count"], r["created_at"]])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=relatorio_admin.csv"})

@app.route("/admin/reports_users.csv")
@admin_required
def admin_reports_users_csv():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    db = get_db()
    rows = db.execute(
        f"SELECT u.id, u.username, COUNT(DISTINCT r.device_name) AS devices, COUNT(r.id) AS registros, COALESCE(SUM(r.fusion_count), 0) AS fusoes "
        f"FROM records r JOIN users u ON u.id = r.user_id {where_sql} GROUP BY u.id, u.username ORDER BY fusoes DESC, devices DESC",
        tuple(params),
    ).fetchall()

    si = StringIO(); writer = csv.writer(si)
    writer.writerow(["user_id", "username", "devices_distintos", "registros", "fusoes"])
    for r in rows:
        writer.writerow([r["id"], r["username"], r["devices"], r["registros"], r["fusoes"]])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=relatorio_por_usuario.csv"})

@app.route("/admin/reports.xlsx")
@admin_required
def admin_reports_xlsx():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str: clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:   clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:   clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    db = get_db()

    rows = db.execute(
        f"SELECT r.id, u.username, r.device_name, r.fusion_count, r.created_at FROM records r JOIN users u ON u.id = r.user_id {where_sql} ORDER BY r.created_at DESC",
        tuple(params),
    ).fetchall()
    devices = db.execute(
        f"SELECT r.device_name, COUNT(*) as registros, SUM(r.fusion_count) as fusoes FROM records r {where_sql} GROUP BY r.device_name ORDER BY fusoes DESC",
        tuple(params),
    ).fetchall()
    users_summary = db.execute(
        f"SELECT u.id, u.username, COUNT(DISTINCT r.device_name) AS devices, COUNT(r.id) AS registros, COALESCE(SUM(r.fusion_count), 0) AS fusoes "
        f"FROM records r JOIN users u ON u.id = r.user_id {where_sql} GROUP BY u.id, u.username ORDER BY fusoes DESC, devices DESC",
        tuple(params),
    ).fetchall()
    by_day = db.execute(
        f"SELECT date(r.created_at) as d, SUM(r.fusion_count) as total FROM records r {where_sql} GROUP BY date(r.created_at) ORDER BY d ASC",
        tuple(params),
    ).fetchall()

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Detalhes"
    ws1.append(["id", "username", "device_name", "fusion_count", "created_at"])
    for r in rows:
        ws1.append([r["id"], r["username"], r["device_name"], r["fusion_count"], r["created_at"]])

    ws2 = wb.create_sheet("Dispositivos")
    ws2.append(["device_name", "registros", "fusoes"])
    for d in devices:
        ws2.append([d["device_name"], d["registros"], d["fusoes"]])

    ws3 = wb.create_sheet("Por Dia")
    ws3.append(["date", "fusoes"])
    total = 0
    for r in by_day:
        v = r["total"] or 0
        total += v
        ws3.append([r["d"], v])
    ws3.append([]); ws3.append(["TOTAL", total])

    ws4 = wb.create_sheet("Por Usuário")
    ws4.append(["user_id", "username", "devices_distintos", "registros", "fusoes"])
    for r in users_summary:
        ws4.append([r["id"], r["username"], r["devices"], r["registros"], r["fusoes"]])

    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return Response(bio.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=relatorio_admin.xlsx"})

# ===== Download de fotos em ZIP por dispositivo =====
@app.route("/admin/photos", methods=["GET", "POST"])
@admin_required
def admin_photos():
    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()
    user_id = request.args.get("user_id", type=int)

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    db = get_db()
    devices = db.execute(
        f"""
        SELECT r.device_name,
               COUNT(DISTINCT r.id) as registros,
               COUNT(p.id) as fotos
        FROM records r
        LEFT JOIN photos p ON p.record_id = r.id
        {where_sql}
        GROUP BY r.device_name
        ORDER BY fotos DESC, registros DESC
        """,
        tuple(params)
    ).fetchall()

    users = db.execute("SELECT id, username FROM users ORDER BY username ASC").fetchall()
    return render_template("admin_photos.html", devices=devices, users=users, selected_user_id=user_id, start=start_str, end=end_str)

@app.route("/admin/photos.zip", methods=["POST"])
@admin_required
def admin_photos_zip():
    start_str = request.form.get("start", "").strip()
    end_str = request.form.get("end", "").strip()
    user_id = request.form.get("user_id", type=int)
    selected = request.form.getlist("devices")

    if not selected:
        flash("Selecione pelo menos um dispositivo.", "error")
        return redirect(url_for("admin_photos", start=start_str, end=end_str, user_id=user_id))

    clauses = []; params = []
    if start_str:
        clauses.append("date(r.created_at) >= date(?)"); params.append(start_str)
    if end_str:
        clauses.append("date(r.created_at) <= date(?)"); params.append(end_str)
    if user_id:
        clauses.append("r.user_id = ?"); params.append(user_id)
    in_clause = " OR ".join(["r.device_name = ?"] * len(selected))
    clauses.append(f"({in_clause})"); params.extend(selected)

    where_sql = "WHERE " + " AND ".join(clauses)

    db = get_db()
    rows = db.execute(
        f"""
        SELECT r.id as record_id, r.device_name, u.username, p.filename
        FROM records r
        JOIN users u ON u.id = r.user_id
        JOIN photos p ON p.record_id = r.id
        {where_sql}
        ORDER BY r.device_name ASC, r.id ASC
        """,
        tuple(params)
    ).fetchall()

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
        base_upload = app.config.get("UPLOAD_FOLDER")
        for row in rows:
            device = row["device_name"]
            rec_id = row["record_id"]
            filename = row["filename"]
            fpath = os.path.join(base_upload, filename)
            if os.path.isfile(fpath):
                arcname = f"{device}/record_{rec_id}/{filename}"
                z.write(fpath, arcname)
    bio.seek(0)
    fname = "fotos_filtradas.zip"
    return Response(
        bio.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# ===== Export do usuário (pessoal) =====
@app.route("/export.csv")
@login_required
def export_csv():
    db = get_db()
    rows = db.execute(
        "SELECT id, device_name, fusion_count, created_at FROM records WHERE user_id = ? ORDER BY created_at DESC",
        (session["user_id"],),
    ).fetchall()
    si = StringIO(); writer = csv.writer(si)
    writer.writerow(["id", "device_name", "fusion_count", "created_at", "photo_urls"])
    for r in rows:
        photos = db.execute("SELECT filename FROM photos WHERE record_id = ?", (r["id"],)).fetchall()
        host = request.host_url.rstrip("/")
        urls = [f"{host}{url_for('uploaded_file', filename=p['filename'])}" for p in photos]
        writer.writerow([r["id"], r["device_name"], r["fusion_count"], r["created_at"], " | ".join(urls)])
    return Response(si.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=registros_splicing.csv"})

# ===== Rota de emergência para resetar senha do admin =====
@app.route("/force_reset_admin")
def force_reset_admin():
    # Controle por variáveis de ambiente
    if os.environ.get("FORCE_RESET_ADMIN", "0") != "1":
        return "Desativado", 403
    token = request.args.get("token", "")
    expected = os.environ.get("RESET_ADMIN_TOKEN", "")
    if not token or token != expected:
        return "Token inválido", 403
    new_pw = os.environ.get("NEW_ADMIN_PASSWORD", "nova123")
    db = get_db()
    # Reseta a senha do primeiro admin encontrado
    row = db.execute("SELECT id FROM users WHERE is_admin = 1 ORDER BY id ASC LIMIT 1").fetchone()
    if not row:
        return "Nenhum admin encontrado", 404
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), row["id"]))
    db.commit()
    return f"Senha do admin (id={row['id']}) resetada para: {new_pw}"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)


@app.route('/admin/workmaps', methods=['GET', 'POST'])
def admin_workmaps():
    if not session.get('is_admin'):
        abort(403)
    with closing(get_db()) as db:
        if request.method == 'POST':
            # Upload a new PDF
            title = request.form.get('title') or 'Mapa de Trabalho'
            file = request.files.get('pdf')
            if not file or not file.filename.lower().endswith('.pdf'):
                flash(('danger', 'Envie um arquivo PDF válido.'))
                return redirect(url_for('admin_workmaps'))
            fname = secure_filename(file.filename)
            os.makedirs(WORKMAP_FOLDER, exist_ok=True)
            dest = os.path.join(WORKMAP_FOLDER, fname)
            file.save(dest)
            db.execute("INSERT INTO work_maps (title, filename) VALUES (?, ?)", (title, fname))
            db.commit()
            flash(('success','Mapa enviado com sucesso.'))
            return redirect(url_for('admin_workmaps'))
        maps = db.execute("SELECT * FROM work_maps ORDER BY uploaded_at DESC").fetchall()
        users = db.execute("SELECT id, username FROM users ORDER BY username").fetchall()
        # get all current grants
        grants = db.execute("SELECT user_id, work_map_id FROM user_work_map_access").fetchall()
        grant_set = {(g['user_id'], g['work_map_id']) for g in grants}
        return render_template('admin_workmaps.html', maps=maps, users=users, grant_set=grant_set)

@app.route('/admin/workmaps/grant', methods=['POST'])
def admin_workmaps_grant():
    if not session.get('is_admin'):
        abort(403)
    user_id = request.form.get('user_id', type=int)
    work_map_id = request.form.get('work_map_id', type=int)  # (opcional)
    action = request.form.get('action','grant')
    with closing(get_db()) as db:
        if action == 'revoke':
            db.execute("DELETE FROM user_work_map_access WHERE user_id=? AND work_map_id=?", (user_id, work_map_id))
        else:
            try:
                db.execute("INSERT OR IGNORE INTO user_work_map_access (user_id, work_map_id) VALUES (?,?)", (user_id, work_map_id))
            except Exception:
                pass
        db.commit()
    flash(('success','Permissões atualizadas.'))
    return redirect(url_for('admin_workmaps'))

@app.route('/workmaps/<int:wm_id>/download')
def workmap_download(wm_id):
    # Admins can download anything; users only if they have access
    with closing(get_db()) as db:
        wm = db.execute("SELECT * FROM work_maps WHERE id=?", (wm_id,)).fetchone()
        if not wm:
            abort(404)
        if not session.get('is_admin'):
            uid = session.get('user_id')
            if not uid or not user_has_access_to_map(uid, wm_id):
                abort(403)
    return send_from_directory(WORKMAP_FOLDER, wm['filename'], as_attachment=True)


@app.route('/records/<int:rec_id>/launch', methods=['POST'])
def record_launch(rec_id):
    uid = session.get('user_id')
    if not uid:
        return redirect(url_for('login'))
    work_map_id = request.form.get('work_map_id', type=int)  # (opcional)
    # Only admin can mark as launched
    if not session.get('is_admin'):
        abort(403)
    with closing(get_db()) as db:
        # Ensure record exists
        rec = db.execute("SELECT * FROM records WHERE id=?", (rec_id,)).fetchone()
        if not rec:
            abort(404)
        # Ensure selected work_map exists
        wm = db.execute("SELECT * FROM work_maps WHERE id=?", (work_map_id,)).fetchone()
        if not wm:
            flash(('danger','Selecione um Mapa de Trabalho válido.'))
            return redirect(url_for('view_record', record_id=rec_id))
        db.execute("UPDATE records SET status='launched', work_map_id=? WHERE id=?", (work_map_id, rec_id))
        db.commit()
    flash(('success','Dispositivo marcado como LANÇADO.'))
    return redirect(url_for('view_record', record_id=rec_id))


@app.route('/my/workmaps')
def my_workmaps():
    uid = session.get('user_id')
    if not uid:
        return redirect(url_for('login'))
    maps = get_user_accessible_maps(uid)
    return render_template('my_workmaps.html', maps=maps)

@app.route("/healthz")
def healthz():
    # basic checks: can open DB and write to DATA_DIR
    ok = True
    checks = {}
    # DB check
    try:
        db = get_db()
        db.execute("SELECT 1").fetchone()
        checks["db"] = "ok"
    except Exception as e:
        checks["db"] = f"error: {e}"
        ok = False
    # Disk check
    if is_writable(DATA_DIR):
        checks["disk"] = "ok"
    else:
        checks["disk"] = "not-writable"
        ok = False
    code = 200 if ok else 503
    try:
        import json
        return app.response_class(json.dumps({"ok": ok, "checks": checks}), status=code, mimetype="application/json")
    except Exception:
        return ("ok" if ok else "not ok", code)



@app.route("/admin/backup")
@login_required
@admin_required
def admin_backup():
    try:
        path = backup_db()
        flash(("success", f"Backup criado: {os.path.basename(path)}"))
    except Exception as e:
        flash(("danger", f"Falha ao criar backup: {e}"))
    return redirect(url_for("admin_home"))



def _schedule_daily_backup():
    # Only start when explicitly enabled
    if os.environ.get("AUTO_BACKUP_DAILY", "0") != "1":
        return
    import threading, datetime, time
    def runner():
        while True:
            now = datetime.datetime.utcnow()
            # Next run at 03:00 UTC
            nxt = now.replace(hour=3, minute=0, second=0, microsecond=0)
            if nxt <= now:
                nxt += datetime.timedelta(days=1)
            sleep_s = (nxt - now).total_seconds()
            try:
                time.sleep(sleep_s)
            except Exception:
                pass
            try:
                backup_db()
            except Exception as e:
                print("Auto-backup error:", e)
    threading.Thread(target=runner, daemon=True).start()

# Kick off auto-backup once at import time
try:
    _schedule_daily_backup()
except Exception as e:
    print("Auto-backup scheduler failed to start:", e)



@app.route("/admin/backups")
@login_required
@admin_required
def admin_backups():
    files = []
    try:
        for name in sorted(os.listdir(BACKUP_DIR)):
            p = os.path.join(BACKUP_DIR, name)
            if os.path.isfile(p) and name.endswith(".db"):
                files.append({
                    "name": name,
                    "size": os.path.getsize(p),
                    "mtime": os.path.getmtime(p),
                })
    except Exception as e:
        flash(("danger", f"Erro ao listar backups: {e}"))
    return render_template("admin_backups.html", files=files)

@app.route("/admin/backups/download/<path:name>")
@login_required
@admin_required
def admin_backup_download(name):
    if "/" in name or "\\" in name or not name.endswith(".db"):
        abort(400)
    return send_from_directory(BACKUP_DIR, name, as_attachment=True)

@app.route("/admin/backups/delete/<path:name>", methods=["POST"])
@login_required
@admin_required
def admin_backup_delete(name):
    if "/" in name or "\\" in name or not name.endswith(".db"):
        abort(400)
    try:
        os.remove(os.path.join(BACKUP_DIR, name))
        flash(("success", f"Backup removido: {name}"))
    except FileNotFoundError:
        flash(("warning", f"Backup não encontrado: {name}"))
    except Exception as e:
        flash(("danger", f"Falha ao remover: {e}"))
    return redirect(url_for("admin_backups"))

@app.template_filter('datetime')
def _fmt_dt(ts):
    try:
        import datetime as _dt
        return _dt.datetime.utcfromtimestamp(int(ts)).strftime('%Y-%m-%d %H:%M UTC')
    except Exception:
        return str(ts)


@app.get("/_debug/db")
def _debug_db():
    from flask import jsonify
    import os
    data_dir = os.getenv("DATA_DIR", "/var/data")
    db_file = os.getenv("DATABASE_FILE", "splice.db")
    db_path = os.path.join(data_dir, db_file)
    try:
        listing = os.listdir(data_dir)
    except Exception as e:
        listing = [f"<error: {e}>"]
    return jsonify({
        "DATA_DIR": data_dir,
        "DATABASE_FILE": db_file,
        "db_path": db_path,
        "db_exists": os.path.exists(db_path),
        "dir_listing": listing,
        "DATABASE_URL": os.getenv("DATABASE_URL")
    })

# === Rota auxiliar /db.json para debug de persistência ===
try:
    from flask import Blueprint, jsonify
    from persist_helper import persist_info
    _persist_bp = Blueprint("persist_debug", __name__)

    @_persist_bp.route("/db.json")
    def _persist_db_json():
        return jsonify(persist_info())

    # Auto registrar no app se existir variável 'app'
    if "app" in globals():
        try:
            app.register_blueprint(_persist_bp)
        except Exception:
            pass
except Exception:
    pass
# === Fim rota auxiliar ===

# --- Endpoints de saúde e debug ---
try:
    from monitor_bp import monitor_bp
    app.register_blueprint(monitor_bp)
except Exception as _bp_e:
    pass

# --- Patch: accept plaintext admin/admin as fallback for convenience ---
def _accept_admin_plaintext(stored_password: str, provided_password: str, username: str) -> bool:
    try:
        if username == "admin" and provided_password == "admin":
            # If DB has plaintext 'admin' or any value, allow admin/admin as a bootstrap login.
            return True
    except Exception:
        pass
    return False


# === devices_feature_patch ===
import sqlite3
from flask import jsonify

def _ensure_devices_schema():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS maps (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        is_active INTEGER DEFAULT 1
    );""")
    cur.execute("""CREATE TABLE IF NOT EXISTS devices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        map_id INTEGER,
        code TEXT,
        address TEXT,
        ports INTEGER,
        feet INTEGER,
        splicer TEXT,
        status TEXT,
        lat REAL,
        lng REAL,
        created_by TEXT,
        updated_by TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT,
        FOREIGN KEY(map_id) REFERENCES maps(id)
    );""")
    cur.execute("""CREATE TABLE IF NOT EXISTS device_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id INTEGER,
        username TEXT,
        UNIQUE(device_id, username)
    );""")
    conn.commit()
    conn.close()

try:
    DB_PATH
except NameError:
    DB_PATH = 'database.db'

try:
    _ensure_devices_schema()
except Exception:
    pass

@app.route('/admin/devices', methods=['GET','POST'])
@login_required
def admin_devices():
    if not is_admin(current_user):
        abort(403)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    if request.method == 'POST':
        code = request.form.get('code')
        map_id = request.form.get('map_id') or None
        address = request.form.get('address')
        ports = request.form.get('ports') or None
        feet = request.form.get('feet') or None
        splicer = request.form.get('splicer')
        status = request.form.get('status')
        lat = request.form.get('lat') or None
        lng = request.form.get('lng') or None
        cur.execute("INSERT INTO devices (map_id, code, address, ports, feet, splicer, status, lat, lng, created_by) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (map_id, code, address, ports, feet, splicer, status, lat, lng, current_user.username))
        conn.commit()
    cur.execute("SELECT id, name FROM maps WHERE is_active=1")
    maps = cur.fetchall()
    cur.execute("SELECT id, code, map_id, address, status FROM devices ORDER BY id DESC")
    devices = cur.fetchall()
    conn.close()
    return render_template('admin_devices.html', maps=maps, devices=devices)

@app.route('/map/<int:map_id>')
@login_required
def map_view(map_id):
    return render_template('map_view.html', map_id=map_id)

@app.route('/api/maps/<int:map_id>/devices')
@login_required
def api_devices_for_map(map_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id, code, address, ports, feet, splicer, status, lat, lng FROM devices WHERE (map_id=? OR ? IS NULL)", (map_id, map_id))
    rows = cur.fetchall()
    conn.close()
    items = []
    for r in rows:
        items.append(dict(id=r[0], code=r[1], address=r[2], ports=r[3], feet=r[4], splicer=r[5], status=r[6], lat=r[7], lng=r[8]))
    return jsonify(items)

@app.route('/device/<int:device_id>', methods=['GET','POST'])
@login_required
def device_edit(device_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    if request.method == 'POST':
        allowed = is_admin(current_user)
        if not allowed:
            cur.execute("SELECT COUNT(*) FROM device_users WHERE device_id=? AND username=?", (device_id, current_user.username))
            allowed = cur.fetchone()[0] > 0
        if not allowed:
            conn.close()
            abort(403)
        fields = ['code','address','ports','feet','splicer','status','lat','lng','map_id']
        updates = ', '.join([f"{f}=?" for f in fields])
        values = [request.form.get(f) for f in fields]
        values.extend([current_user.username, device_id])
        cur.execute(f"UPDATE devices SET {updates}, updated_by=?, updated_at=datetime('now') WHERE id=?", values)
        conn.commit()
    cur.execute("SELECT id, map_id, code, address, ports, feet, splicer, status, lat, lng FROM devices WHERE id=?", (device_id,))
    row = cur.fetchone()
    cur.execute("SELECT id, name FROM maps WHERE is_active=1")
    maps = cur.fetchall()
    cur.execute("SELECT username FROM device_users WHERE device_id=?", (device_id,))
    assigned = [r[0] for r in cur.fetchall()]
    conn.close()
    if not row:
        abort(404)
    device = dict(id=row[0], map_id=row[1], code=row[2], address=row[3], ports=row[4], feet=row[5], splicer=row[6], status=row[7], lat=row[8], lng=row[9])
        # fetch tasks for this device
    cur = sqlite3.connect(DB_PATH).cursor()
    cur.execute("SELECT id, task_type, notes, status FROM device_tasks WHERE device_id=? ORDER BY id DESC", (device_id,))
    tasks = cur.fetchall()
        cur = sqlite3.connect(DB_PATH).cursor()
    cur.execute("SELECT id, filename, caption FROM device_photos WHERE device_id=? ORDER BY id DESC", (device_id,))
    photos = cur.fetchall()
    return render_template('device_edit.html', device=device, maps=maps, assigned=assigned, tasks=tasks, photos=photos)

@app.route('/admin/devices/<int:device_id>/assign', methods=['POST'])
@login_required
def assign_user_to_device(device_id):
    if not is_admin(current_user):
        abort(403)
    username = request.form.get('username')
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("INSERT OR IGNORE INTO device_users (device_id, username) VALUES (?,?)", (device_id, username))
    conn.commit()
    conn.close()
    return redirect(url_for('device_edit', device_id=device_id))


# === maps_and_tasks_patch ===
import csv
from io import TextIOWrapper
from flask import send_file

def _ensure_tasks_schema():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS device_tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id INTEGER NOT NULL,
        task_type TEXT NOT NULL,  -- 'Splice','Tests','Place','Troubleshooting'
        notes TEXT,
        status TEXT DEFAULT 'pending', -- 'pending' or 'done'
        created_by TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );""")
    conn.commit()
    conn.close()

try:
    _ensure_tasks_schema()
except Exception:
    pass

# List maps (id, name) and link to /map/<id>
@app.route('/maps')
@login_required
def list_maps():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id, name, description FROM maps WHERE is_active=1 ORDER BY id DESC")
    maps = cur.fetchall()
    conn.close()
    return render_template('maps.html', maps=maps)

# Device tasks view + quick add from device page
@app.route('/device/<int:device_id>/tasks', methods=['POST'])
@login_required
def device_add_task(device_id):
    # Anyone with edit permission (or admin) can add task
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    allowed = 1 if is_admin(current_user) else 0
    if not allowed:
        cur.execute("SELECT COUNT(*) FROM device_users WHERE device_id=? AND username=?", (device_id, current_user.username))
        allowed = 1 if cur.fetchone()[0] > 0 else 0
    if not allowed:
        conn.close()
        abort(403)
    task_type = request.form.get('task_type')
    notes = request.form.get('notes')
    cur.execute("INSERT INTO device_tasks (device_id, task_type, notes, created_by) VALUES (?,?,?,?)",
                (device_id, task_type, notes, current_user.username))
    conn.commit()
    conn.close()
    return redirect(url_for('device_edit', device_id=device_id))

@app.route('/device/<int:device_id>/tasks/<int:task_id>/toggle', methods=['POST'])
@login_required
def device_toggle_task(device_id, task_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # permission
    allowed = 1 if is_admin(current_user) else 0
    if not allowed:
        cur.execute("SELECT COUNT(*) FROM device_users WHERE device_id=? AND username=?", (device_id, current_user.username))
        allowed = 1 if cur.fetchone()[0] > 0 else 0
    if not allowed:
        conn.close()
        abort(403)
    # toggle
    cur.execute("SELECT status FROM device_tasks WHERE id=? AND device_id=?", (task_id, device_id))
    row = cur.fetchone()
    if row:
        new_status = 'done' if row[0] != 'done' else 'pending'
        cur.execute("UPDATE device_tasks SET status=? WHERE id=?", (new_status, task_id))
        conn.commit()
    conn.close()
    return redirect(url_for('device_edit', device_id=device_id))

# CSV import on admin devices
@app.route('/admin/devices/import', methods=['POST'])
@login_required
def admin_devices_import():
    if not is_admin(current_user):
        abort(403)
    if 'file' not in request.files:
        flash('Selecione um CSV', 'warning')
        return redirect(url_for('admin_devices'))
    f = request.files['file']
    stream = TextIOWrapper(f.stream, encoding='utf-8')
    reader = csv.DictReader(stream)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    count = 0
    for row in reader:
        code = row.get('code') or row.get('codigo') or row.get('device')
        address = row.get('address')
        ports = row.get('ports')
        feet = row.get('feet')
        splicer = row.get('splicer')
        status = row.get('status')
        lat = row.get('lat') or row.get('latitude')
        lng = row.get('lng') or row.get('longitude')
        map_id = row.get('map_id') or row.get('map')
        cur.execute("INSERT INTO devices (map_id, code, address, ports, feet, splicer, status, lat, lng, created_by) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (map_id, code, address, ports, feet, splicer, status, lat, lng, current_user.username))
        count += 1
    conn.commit()
    conn.close()
    flash(f'Importados {count} dispositivos do CSV.', 'success')
    return redirect(url_for('admin_devices'))


# === maps_photos_reports_patch ===
import os, csv, sqlite3, secrets
from flask import send_file, make_response
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'png','jpg','jpeg','gif','webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def _ensure_maps_schema():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS maps (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        is_active INTEGER DEFAULT 1
    );""")
    conn.commit(); conn.close()

def _ensure_photos_schema():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS device_photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        caption TEXT,
        created_by TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );""")
    conn.commit(); conn.close()

try:
    _ensure_maps_schema()
    _ensure_photos_schema()
except Exception:
    pass

# --- Admin maps create/list ---
@app.route('/admin/maps', methods=['GET','POST'])
@login_required
def admin_maps():
    if not is_admin(current_user):
        abort(403)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        cur.execute("INSERT INTO maps (name, description, is_active) VALUES (?,?,1)", (name, description))
        conn.commit()
    cur.execute("SELECT id, name, description, is_active FROM maps ORDER BY id DESC")
    maps = cur.fetchall()
    conn.close()
    return render_template('admin_maps.html', maps=maps)

# --- Device photos upload (max 6) ---
@app.route('/device/<int:device_id>/upload', methods=['POST'])
@login_required
def device_upload_photos(device_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # permission: admin or assigned
    allowed = 1 if is_admin(current_user) else 0
    if not allowed:
        cur.execute("SELECT COUNT(*) FROM device_users WHERE device_id=? AND username=?", (device_id, current_user.username))
        allowed = 1 if cur.fetchone()[0] > 0 else 0
    if not allowed:
        conn.close(); abort(403)

    files = request.files.getlist('photos')
    captions = request.form.getlist('captions')
    # count existing
    cur.execute("SELECT COUNT(*) FROM device_photos WHERE device_id=?", (device_id,))
    existing = cur.fetchone()[0]
    remaining = max(0, 6 - existing)
    saved = 0
    for i, f in enumerate(files[:remaining]):
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            ext = filename.rsplit('.',1)[1].lower()
            newname = f"{device_id}_{secrets.token_hex(6)}.{ext}"
            path = os.path.join(UPLOAD_FOLDER, newname)
            f.save(path)
            caption = captions[i] if i < len(captions) else None
            cur.execute("INSERT INTO device_photos (device_id, filename, caption, created_by) VALUES (?,?,?,?)",
                        (device_id, newname, caption, current_user.username))
            saved += 1
    conn.commit(); conn.close()
    flash(f'Upload concluído: {saved} arquivo(s).', 'success')
    return redirect(url_for('device_edit', device_id=device_id))

@app.route('/device/<int:device_id>/photo/<int:photo_id>/delete', methods=['POST'])
@login_required
def device_delete_photo(device_id, photo_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    allowed = 1 if is_admin(current_user) else 0
    if not allowed:
        cur.execute("SELECT COUNT(*) FROM device_users WHERE device_id=? AND username=?", (device_id, current_user.username))
        allowed = 1 if cur.fetchone()[0] > 0 else 0
    if not allowed:
        conn.close(); abort(403)
    cur.execute("SELECT filename FROM device_photos WHERE id=? AND device_id=?", (photo_id, device_id))
    row = cur.fetchone()
    if row:
        filepath = os.path.join(UPLOAD_FOLDER, row[0])
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception:
            pass
        cur.execute("DELETE FROM device_photos WHERE id=?", (photo_id,))
        conn.commit()
    conn.close()
    return redirect(url_for('device_edit', device_id=device_id))

# --- Reports by map ---
@app.route('/map/<int:map_id>/report')
@login_required
def map_report(map_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id, name, description FROM maps WHERE id=?", (map_id,))
    map_row = cur.fetchone()
    cur.execute("SELECT id, code, address, ports, feet, splicer, status, lat, lng FROM devices WHERE (map_id=? OR ? IS NULL) ORDER BY id", (map_id, map_id))
    devices = cur.fetchall()
    # totals
    total = len(devices)
    done = 0
    for d in devices:
        st = (d[6] or '').lower()
        if st in ('feito','ok','done','completed','concluido','concluído'):
            done += 1
    pend = total - done
    conn.close()
    return render_template('map_report.html', map_row=map_row, devices=devices, total=total, done=done, pend=pend)

@app.route('/map/<int:map_id>/report.csv')
@login_required
def map_report_csv(map_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM maps WHERE id=?", (map_id,))
    m = cur.fetchone()
    cur.execute("SELECT id, code, address, ports, feet, splicer, status, lat, lng FROM devices WHERE (map_id=? OR ? IS NULL) ORDER BY id", (map_id, map_id))
    rows = cur.fetchall()
    conn.close()
    headers = ['id','code','address','ports','feet','splicer','status','lat','lng']
    import io
    si = io.StringIO()
    si.write('map_id,map_name\n')
    si.write(f"{map_id},{(m[1] if m else '')}\n\n")
    si.write(','.join(headers)+'\n')
    for r in rows:
        line = [str(r[0]), r[1] or '', r[2] or '', str(r[3] or ''), str(r[4] or ''), r[5] or '', r[6] or '', str(r[7] or ''), str(r[8] or '')]
        si.write(','.join([c.replace(',', ' ') for c in line]) + '\n')
    output = make_response(si.getvalue())
    output.headers['Content-Disposition'] = f'attachment; filename=map_{map_id}_report.csv'
    output.headers['Content-Type'] = 'text/csv'
    return output
